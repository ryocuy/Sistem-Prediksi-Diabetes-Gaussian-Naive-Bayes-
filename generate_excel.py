import csv
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hitungan"

# 1. Tulis Dataset ke Kolom A-E
with open("diabetes.csv", "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, 1):
        for col_idx, val in enumerate(row, 1):
            try:
                if "." in val:
                    val = float(val)
                else:
                    val = int(val)
            except ValueError:
                pass
            ws.cell(row=row_idx, column=col_idx, value=val)

# 2. Setup Layout Hitungan di sebelah kanan
ws['G1'] = "Naive Bayes Gaussian - Dataset Diabetes"

ws['G3'] = "Kelas"
ws['H3'] = "Jumlah"
ws['I3'] = "P(H)"
ws['G4'] = 1; ws['H4'] = 50; ws['I4'] = "=H4/100"
ws['G5'] = 2; ws['H5'] = 50; ws['I5'] = "=H5/100"

ws['G7'] = "Kelas 1"
ws['G8'] = "Rata-rata"; ws['H8'] = "Age"; ws['I8'] = "BMI"; ws['J8'] = "Glucose"; ws['K8'] = "Insulin"
ws['H9'] = "=AVERAGE(A2:A51)"; ws['I9'] = "=AVERAGE(B2:B51)"; ws['J9'] = "=AVERAGE(C2:C51)"; ws['K9'] = "=AVERAGE(D2:D51)"
ws['G10'] = "Varian"; ws['H10'] = "=VAR.S(A2:A51)"; ws['I10'] = "=VAR.S(B2:B51)"; ws['J10'] = "=VAR.S(C2:C51)"; ws['K10'] = "=VAR.S(D2:D51)"

ws['G12'] = "Kelas 2"
ws['G13'] = "Rata-rata"; ws['H13'] = "=AVERAGE(A52:A101)"; ws['I13'] = "=AVERAGE(B52:B101)"; ws['J13'] = "=AVERAGE(C52:C101)"; ws['K13'] = "=AVERAGE(D52:D101)"
ws['G14'] = "Varian"; ws['H14'] = "=VAR.S(A52:A101)"; ws['I14'] = "=VAR.S(B52:B101)"; ws['J14'] = "=VAR.S(C52:C101)"; ws['K14'] = "=VAR.S(D52:D101)"

ws['H16'] = "Nilai"; ws['I16'] = "Kelas 1"; ws['J16'] = "Kelas 2"

# FIXED: Use -1 * (x - mean)^2 to avoid Excel unary minus bug
ws['G17'] = "Age"; ws['H17'] = 30
ws['I17'] = "=1/SQRT(2*3.14*H10)*(2.72^(-1*((H17-H9)^2)/(2*H10)))"
ws['J17'] = "=1/SQRT(2*3.14*H14)*(2.72^(-1*((H17-H13)^2)/(2*H14)))"

ws['G18'] = "BMI"; ws['H18'] = 25
ws['I18'] = "=1/SQRT(2*3.14*I10)*(2.72^(-1*((H18-I9)^2)/(2*I10)))"
ws['J18'] = "=1/SQRT(2*3.14*I14)*(2.72^(-1*((H18-I13)^2)/(2*I14)))"

ws['G19'] = "Glucose"; ws['H19'] = 120
ws['I19'] = "=1/SQRT(2*3.14*J10)*(2.72^(-1*((H19-J9)^2)/(2*J10)))"
ws['J19'] = "=1/SQRT(2*3.14*J14)*(2.72^(-1*((H19-J13)^2)/(2*J14)))"

ws['G20'] = "Insulin"; ws['H20'] = 5.0
ws['I20'] = "=1/SQRT(2*3.14*K10)*(2.72^(-1*((H20-K9)^2)/(2*K10)))"
ws['J20'] = "=1/SQRT(2*3.14*K14)*(2.72^(-1*((H20-K13)^2)/(2*K14)))"

ws['G21'] = "output?"; ws['H21'] = 1
ws['G22'] = "Akhir Kali"; ws['H22'] = "Kelas 1"; ws['I22'] = "Kelas 2"
ws['H23'] = "=I4*I17*I18*I19*I20"
ws['I23'] = "=I5*J17*J18*J19*J20"

ws['G25'] = "Hasil Prediksi"
ws['H25'] = '=IF(H23>I23, "Kelas 1", "Kelas 2")'
ws['G26'] = "cari nilai probabilitas tertinggi"
ws['H26'] = '=IF(H23>I23, "Kelas 1 (Sehat)", "Kelas 2 (Sakit)")'

wb.save("Hitungan_UAS_Naive_Bayes.xlsx")
print("File Hitungan_UAS_Naive_Bayes.xlsx diperbaiki!")
